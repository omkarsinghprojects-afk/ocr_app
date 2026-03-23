"""
OCR Web App — FastAPI Backend (Gemini API edition — FREE)
==========================================================
- Upload PDF / PNG / JPG / JPEG / TIFF / BMP / WEBP from any phone or browser
- OCR via Google Gemini API (gemini-1.5-flash) — FREE tier: 1500 requests/day
- All results saved to SQLite
- Download all results as Excel anytime

Deploy on Render / Railway:
  Set env vars:  GEMINI_API_KEY  and  APP_PASSWORD
"""

import os

APP_PASSWORD   = os.environ.get("APP_PASSWORD", "ocr1234")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
DPI            = 150
UPLOAD_FOLDER  = "uploads"
DB_FILE        = "ocr_results.db"

import base64
import io
import json
import re
import sqlite3
import uuid
from datetime import datetime
from pathlib import Path

import uvicorn
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse

try:
    from pdf2image import convert_from_path
except ImportError:
    print("ERROR: pip install pdf2image"); exit(1)

try:
    import google.generativeai as genai
except ImportError:
    print("ERROR: pip install google-generativeai"); exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl"); exit(1)

from PIL import Image

SUPPORTED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}

app = FastAPI(title="OCR Form Scanner")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
Path(UPLOAD_FOLDER).mkdir(exist_ok=True)

def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS scans (
            id TEXT PRIMARY KEY, filename TEXT, filetype TEXT,
            scanned_at TEXT, fields_json TEXT
        )
    """)
    conn.commit(); conn.close()

init_db()

EXTRACT_PROMPT = """You are an expert data entry operator reading a scanned application form.
Extract EVERY SINGLE field visible on this form — go top to bottom, left to right.
Include ALL of: personal details, contact info, address, identity documents (Aadhaar, PAN, Voter ID),
education, employment, bank details, physical details, declarations, signatures, checkboxes — everything.
Rules:
- Every value must be a plain string. No nested objects. No arrays.
- Blank fields: write "" as the value. Do NOT skip them.
- Unclear handwriting: make your best guess and include it.
- Return ONLY a flat JSON object. No explanation. No markdown. No code fences.
Example: {"Full Name": "John Smith", "Date of Birth": "15/08/1990", "Mobile": "9876543210"}
"""

def image_to_bytes(image: Image.Image, max_size: int = 1600) -> bytes:
    w, h = image.size
    if max(w, h) > max_size:
        scale = max_size / max(w, h)
        image = image.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    buf = io.BytesIO()
    image.convert("RGB").save(buf, format="JPEG", quality=88)
    return buf.getvalue()

def gemini_ocr(image: Image.Image) -> dict:
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel("gemini-1.5-flash-002")
    img_part = {"mime_type": "image/jpeg", "data": image_to_bytes(image)}
    response = model.generate_content(
        [EXTRACT_PROMPT, img_part],
        generation_config={"temperature": 0.1, "max_output_tokens": 4096},
    )
    return parse_json(response.text)

def parse_json(raw: str) -> dict:
    if not raw: return {}
    raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()
    m = re.search(r"\{.*\}", raw, re.DOTALL)
    if m:
        try: return json.loads(m.group(0))
        except json.JSONDecodeError: pass
    result = {}
    for line in raw.splitlines():
        if ":" in line:
            k, _, v = line.partition(":")
            k = k.strip().strip('"').strip("'")
            v = v.strip().strip('",').strip("'")
            if k: result[k] = v
    return result

def to_safe(value) -> str:
    if value is None: return ""
    if isinstance(value, dict): return ", ".join(f"{k}: {v}" for k, v in value.items())
    if isinstance(value, list): return " | ".join(to_safe(i) for i in value)
    return str(value)

def file_to_images(path: str) -> list:
    ext = Path(path).suffix.lower()
    if ext == ".pdf": return convert_from_path(path, dpi=DPI)
    img = Image.open(path)
    if hasattr(img, "n_frames") and img.n_frames > 1:
        frames = []
        for i in range(img.n_frames):
            img.seek(i); frames.append(img.copy().convert("RGB"))
        return frames
    return [img.convert("RGB")]

def merge_pages(pages: list) -> dict:
    merged = {}
    for d in pages:
        for k, v in d.items():
            if not v: continue
            if k not in merged or not merged[k]: merged[k] = v
            elif merged[k] != v: merged[k] = f"{to_safe(merged[k])} | {to_safe(v)}"
    return merged

def build_excel(rows: list) -> bytes:
    if not rows: return b""
    all_keys = ["File Name", "File Type", "Scanned At"]
    seen = set(all_keys)
    for row in rows:
        for k in json.loads(row["fields_json"]):
            if k not in seen: all_keys.append(k); seen.add(k)
    wb = Workbook(); ws = wb.active; ws.title = "Application Data"; ws.freeze_panes = "A2"
    HFILL  = PatternFill("solid", fgColor="1F4E79")
    HFONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    AFILL  = PatternFill("solid", fgColor="D6E4F0")
    NFILL  = PatternFill("solid", fgColor="FFFFFF")
    THIN   = Side(style="thin", color="B0C4DE")
    BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for ci, key in enumerate(all_keys, 1):
        c = ws.cell(row=1, column=ci, value=key)
        c.font = HFONT; c.fill = HFILL; c.alignment = CENTER; c.border = BDR
    for ri, row in enumerate(rows, 2):
        fields = json.loads(row["fields_json"]); fill = AFILL if ri % 2 == 0 else NFILL
        ws.row_dimensions[ri].height = 20
        for ci, key in enumerate(all_keys, 1):
            val = row["filename"] if key == "File Name" else row["filetype"] if key == "File Type" else row["scanned_at"] if key == "Scanned At" else to_safe(fields.get(key, ""))
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=9); c.fill = fill; c.alignment = LEFT; c.border = BDR
    for ci, key in enumerate(all_keys, 1):
        mx = len(key)
        for row in rows:
            v = to_safe(json.loads(row["fields_json"]).get(key, ""))
            mx = max(mx, min(len(v), 50))
        ws.column_dimensions[get_column_letter(ci)].width = min(mx + 4, 52)
    ws2 = wb.create_sheet("Summary")
    for r, (lbl, val) in enumerate([("Total scans", len(rows)), ("Unique fields", len(all_keys)-3), ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M"))], 1):
        ws2.cell(row=r, column=1, value=lbl).font = Font(name="Arial", bold=True)
        ws2.cell(row=r, column=2, value=val).font = Font(name="Arial")
    ws2.column_dimensions["A"].width = 20; ws2.column_dimensions["B"].width = 30
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

@app.get("/", response_class=HTMLResponse)
async def home():
    with open("templates/index.html") as f: return f.read()

@app.post("/login")
async def login(password: str = Form(...)):
    if password != APP_PASSWORD: raise HTTPException(status_code=401, detail="Wrong password")
    return {"status": "ok"}

@app.post("/upload")
async def upload(password: str = Form(...), file: UploadFile = File(...)):
    if password != APP_PASSWORD: raise HTTPException(status_code=401, detail="Wrong password")
    if not GEMINI_API_KEY: raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured on server")
    ext = Path(file.filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS: raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")
    tmp_path = f"{UPLOAD_FOLDER}/{uuid.uuid4()}{ext}"
    with open(tmp_path, "wb") as f_out: f_out.write(await file.read())
    try:
        images       = file_to_images(tmp_path)
        page_results = [gemini_ocr(img) for img in images]
        merged       = merge_pages(page_results)
    finally:
        os.remove(tmp_path)
    scan_id = str(uuid.uuid4()); conn = get_db()
    conn.execute("INSERT INTO scans (id, filename, filetype, scanned_at, fields_json) VALUES (?,?,?,?,?)",
        (scan_id, file.filename, ext.lstrip("."), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), json.dumps(merged)))
    conn.commit(); conn.close()
    return {"status": "ok", "scan_id": scan_id, "filename": file.filename, "fields": merged, "count": len(merged)}

@app.get("/scans")
async def list_scans(password: str):
    if password != APP_PASSWORD: raise HTTPException(status_code=401, detail="Wrong password")
    conn = get_db()
    rows = conn.execute("SELECT id, filename, filetype, scanned_at FROM scans ORDER BY scanned_at DESC").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.get("/download-excel")
async def download_excel(password: str):
    if password != APP_PASSWORD: raise HTTPException(status_code=401, detail="Wrong password")
    conn = get_db()
    rows = conn.execute("SELECT * FROM scans ORDER BY scanned_at ASC").fetchall()
    conn.close()
    if not rows: raise HTTPException(status_code=404, detail="No scans yet")
    xlsx  = build_excel([dict(r) for r in rows])
    fname = f"scans_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.delete("/clear")
async def clear_all(password: str):
    if password != APP_PASSWORD: raise HTTPException(status_code=401, detail="Wrong password")
    conn = get_db(); conn.execute("DELETE FROM scans"); conn.commit(); conn.close()
    return {"status": "cleared"}

if __name__ == "__main__":
    print(f"\n{'='*50}\n  OCR Web App — http://localhost:8000\n  Password: {APP_PASSWORD}\n{'='*50}\n")
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=False)
